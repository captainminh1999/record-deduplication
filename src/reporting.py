"""Step 5 of 6: Reporting (Review and Export)

Generates an Excel workbook listing high-confidence duplicate pairs side by side for human review. See README for details.
"""

from __future__ import annotations

import click
import os
import time
import pandas as pd
from openpyxl.styles import PatternFill
from pandas import ExcelWriter
from .utils import log_run, LOG_PATH


def main(
    dupes_path: str = "data/outputs/high_confidence.csv",
    cleaned_path: str = "data/outputs/cleaned.csv",
    report_path: str = "data/outputs/manual_review.xlsx",
    log_path: str = LOG_PATH,
    gpt_review_path: str = "data/outputs/gpt_review.json",
) -> None:
    """Create a merge suggestion workbook, now with GPT review results."""

    print("📊 Starting Excel report generation...")
    
    start_time = time.time()
    report_stats = {
        "input_stats": {},
        "probability_bands": {},
        "gpt_review": {"available": False},
        "output_sheets": {}
    }

    # Load and track input data stats
    dupes = pd.read_csv(dupes_path)
    cleaned = pd.read_csv(cleaned_path).set_index("record_id")
    report_stats["input_stats"] = {
        "total_pairs": len(dupes),
        "unique_records": len(cleaned),
        "mean_probability": float(dupes["prob"].mean()),
        "probability_quantiles": {
            "25%": float(dupes["prob"].quantile(0.25)),
            "50%": float(dupes["prob"].quantile(0.50)),
            "75%": float(dupes["prob"].quantile(0.75)),
            "90%": float(dupes["prob"].quantile(0.90))
        }
    }

    # Combine duplicate pairs with full record details from both sides.
    left = cleaned.loc[dupes["record_id_1"]].add_suffix("_1").reset_index()
    right = cleaned.loc[dupes["record_id_2"]].add_suffix("_2").reset_index()
    merged = pd.concat([dupes.reset_index(drop=True), left, right], axis=1)

    # Track probability band statistics
    prob_bands = {
        "high_confidence": merged[merged["prob"] >= 0.9],
        "manual_review": merged[(merged["prob"] >= 0.6) & (merged["prob"] < 0.9)],
        "low_confidence": merged[merged["prob"] < 0.6]
    }
    
    report_stats["probability_bands"] = {
        band: {
            "count": len(df),
            "mean_probability": float(df["prob"].mean()) if len(df) > 0 else 0.0
        }
        for band, df in prob_bands.items()
    }

    # Load and process GPT review results if available
    gpt_flat = []
    if os.path.exists(gpt_review_path):
        import json
        
        with open(gpt_review_path, "r", encoding="utf-8") as f:
            gpt_review = json.load(f)
            
        # Track GPT review statistics
        total_clusters = len(gpt_review)
        processed_clusters = 0
        total_groups = 0
        total_records = 0
        confidence_sum = 0.0
        
        for entry in gpt_review:
            cluster_id = entry.get("cluster_id")
            record_ids = entry.get("record_ids", [])
            canonical_groups = entry.get("canonical_groups", [])
            
            # Only process clusters with non-empty canonical_groups
            if canonical_groups and isinstance(canonical_groups, list) and len(canonical_groups) > 0:
                processed_clusters += 1
                total_groups += len(canonical_groups)
                
                for group in canonical_groups:
                    # Required fields
                    primary_organization = group.get("primary_organization")
                    canonical_domains = group.get("canonical_domains", [])
                    canonical_record_ids = group.get("record_ids", [])
                    confidence = group.get("confidence", 0.0)
                    
                    if primary_organization and canonical_record_ids:
                        total_records += len(canonical_record_ids)
                        confidence_sum += confidence
                        gpt_flat.append({
                            "cluster_id": cluster_id,
                            "primary_organization": primary_organization,
                            "canonical_domains": canonical_domains,
                            "record_ids": canonical_record_ids,
                            "confidence": confidence
                        })
        
        # Update GPT review stats
        report_stats["gpt_review"] = {
            "available": True,
            "total_clusters": total_clusters,
            "processed_clusters": processed_clusters,
            "total_groups": total_groups,  # Fixed: changed from total_canonical_groups
            "total_records_processed": total_records,
            "mean_confidence": float(confidence_sum / len(gpt_flat)) if gpt_flat else 0.0,
            "records_per_group": float(total_records / total_groups) if total_groups > 0 else 0.0
        }

    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with ExcelWriter(report_path, engine="openpyxl") as writer:
        prob_bands["high_confidence"].to_excel(writer, sheet_name="high_confidence", index=False)
        prob_bands["manual_review"].to_excel(writer, sheet_name="manual_review", index=False)
        if gpt_flat:
            pd.DataFrame(gpt_flat).to_excel(writer, sheet_name="gpt_review", index=False)

        # Optional formatting highlighting borderline pairs for review.
        workbook = writer.book
        review_ws = writer.sheets["manual_review"]
        fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        prob_idx = prob_bands["manual_review"].columns.get_loc("prob")
        if not isinstance(prob_idx, int):
            raise TypeError("Expected a unique 'prob' column")
        score_col = prob_idx + 1
        for row in review_ws.iter_rows(
            min_row=2, max_row=review_ws.max_row, min_col=score_col, max_col=score_col
        ):
            for cell in row:
                cell.fill = fill

    # Print comprehensive terminal output
    print(f"\n📊 Excel Report Generation Complete!")
    print(f"─" * 50)
    print(f"📊 Data Overview:")
    print(f"  • Total pairs analyzed:  {report_stats['input_stats']['total_pairs']:,}")
    print(f"  • Unique records:        {report_stats['input_stats']['unique_records']:,}")
    print(f"  • Mean probability:      {report_stats['input_stats']['mean_probability']:.3f}")
    
    print(f"\n📈 Probability Distribution:")
    quants = report_stats['input_stats']['probability_quantiles']
    print(f"  • 25th percentile:       {quants['25%']:.3f}")
    print(f"  • 50th percentile:       {quants['50%']:.3f}")
    print(f"  • 75th percentile:       {quants['75%']:.3f}")
    print(f"  • 90th percentile:       {quants['90%']:.3f}")
    
    print(f"\n📋 Excel Sheets Created:")
    for sheet_name, count in report_stats['output_sheets'].items():
        print(f"  • {sheet_name:<15} {count:,} pairs")
    
    if report_stats['gpt_review']['available']:
        gpt_stats = report_stats['gpt_review']
        print(f"\n🤖 GPT Review Results:")
        try:
            print(f"  • Total groups:          {gpt_stats.get('total_groups', 0):,}")
            print(f"  • Total records:         {gpt_stats.get('total_records_processed', 0):,}")
            print(f"  • Records per group:     {gpt_stats.get('records_per_group', 0.0):.1f}")
            if 'processed_clusters' in gpt_stats:
                print(f"  • Processed clusters:    {gpt_stats['processed_clusters']:,}")
        except Exception as e:
            print(f"  • Error displaying GPT stats: {e}")
    else:
        print(f"\n🤖 GPT Review: Not available (no GPT analysis found)")
    
    total_pairs = len(prob_bands['high_confidence']) + len(prob_bands['manual_review'])
    print(f"\n💾 Files Created:")
    print(f"  • Excel workbook:        {report_path}")
    print(f"  • Total pairs for review: {total_pairs:,}")
    
    if len(prob_bands['high_confidence']) > 0:
        print(f"\n✅ Success! Found {len(prob_bands['high_confidence']):,} high-confidence duplicate pairs")
        print(f"   Review the 'high_confidence' sheet first")
    
    if len(prob_bands['manual_review']) > 0:
        print(f"   Also review {len(prob_bands['manual_review']):,} borderline pairs in 'manual_review' sheet")
    
    if total_pairs == 0:
        print(f"\n⚠️  No duplicate pairs found for review")
        print(f"   Consider lowering confidence thresholds or improving training data")
    
    print(f"\n📖 How to Review:")
    print(f"   1. Open {report_path} in Excel")
    print(f"   2. Check 'high_confidence' sheet for likely duplicates")
    print(f"   3. Review 'manual_review' sheet for borderline cases")
    if report_stats['gpt_review']['available']:
        print(f"   4. Reference 'gpt_review' sheet for AI analysis")

    end_time = time.time()
    log_run(
        "reporting",
        start_time,
        end_time,
        len(prob_bands["high_confidence"]) + len(prob_bands["manual_review"]),
        log_path=log_path,
    )


@click.command()
@click.option("--duplicates-path", default="data/outputs/high_confidence.csv", show_default=True)
@click.option("--cleaned-path", default="data/outputs/cleaned.csv", show_default=True)
@click.option("--report-path", default="data/outputs/manual_review.xlsx", show_default=True)
@click.option("--log-path", default=LOG_PATH, show_default=True)
@click.option("--gpt-review-path", default="data/outputs/gpt_review.json", show_default=True, help="Path to GPT review JSON file.")
def cli(duplicates_path: str, cleaned_path: str, report_path: str, log_path: str, gpt_review_path: str) -> None:
    """CLI wrapper for :func:`main`."""

    main(duplicates_path, cleaned_path, report_path, log_path, gpt_review_path)


if __name__ == "__main__":  # pragma: no cover - sanity run
    print("\u23e9 started reporting")
    cli()
