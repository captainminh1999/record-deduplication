import os
import pandas as pd
import tempfile
import unittest
from openpyxl import load_workbook

from src import reporting


class ReportingTest(unittest.TestCase):
    def test_reporting_writes_excel(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            dupes_path = os.path.join(tmpdir, "dupes.csv")
            cleaned_path = os.path.join(tmpdir, "cleaned.csv")
            report_path = os.path.join(tmpdir, "manual_review.xlsx")

            pd.DataFrame({
                "record_id_1": [1, 2],
                "record_id_2": [2, 3],
                "prob": [0.95, 0.65],
            }).to_csv(dupes_path, index=False)

            pd.DataFrame({
                "record_id": [1, 2, 3],
                "company_clean": ["a", "b", "c"],
                "domain_clean": ["a.com", "b.com", "c.com"],
                "phone_clean": ["1", "2", "3"],
                "address_clean": ["a", "b", "c"],
            }).to_csv(cleaned_path, index=False)

            reporting.main(dupes_path, cleaned_path, report_path)

            self.assertTrue(os.path.exists(report_path))
            wb = load_workbook(report_path)
            self.assertIn("manual_review", wb.sheetnames)
            self.assertIn("high_confidence", wb.sheetnames)

    def test_reporting_creates_dirs(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            dupes_path = os.path.join(tmpdir, "dupes.csv")
            cleaned_path = os.path.join(tmpdir, "cleaned.csv")
            report_path = os.path.join(tmpdir, "reports", "manual_review.xlsx")

            pd.DataFrame({
                "record_id_1": [1],
                "record_id_2": [2],
                "prob": [0.95],
            }).to_csv(dupes_path, index=False)

            pd.DataFrame({
                "record_id": [1, 2],
                "company_clean": ["a", "b"],
                "domain_clean": ["a.com", "b.com"],
                "phone_clean": ["1", "2"],
                "address_clean": ["a", "b"],
            }).to_csv(cleaned_path, index=False)

            reporting.main(dupes_path, cleaned_path, report_path)

            self.assertTrue(os.path.exists(report_path))


if __name__ == "__main__":
    unittest.main()
