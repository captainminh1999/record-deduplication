"""Reporting Engine - Business Logic

Core reporting functionality for the record deduplication pipeline.
This module contains the main report generation algorithms and business logic,
separated from CLI and formatting concerns.
"""

from __future__ import annotations

import os
import json
from typing import Dict, List, Tuple, Any, Optional

import pandas as pd
from openpyxl.styles import PatternFill
from pandas import ExcelWriter


class ReportingEngine:
    """Core reporting engine for record deduplication."""
    
    def __init__(self):
        self.report_stats = {}
    
    def generate_report(
        self,
        dupes_path: str,
        cleaned_path: str,
        report_path: str,
        gpt_review_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Generate comprehensive Excel report from duplicate pairs and GPT analysis.
        
        Args:
            dupes_path: Path to high-confidence duplicates CSV
            cleaned_path: Path to cleaned records CSV
            report_path: Path to save Excel report
            gpt_review_path: Path to GPT review JSON file (optional)
            
        Returns:
            Dictionary containing report statistics and metadata
        """
        # Initialize report statistics
        self.report_stats = {
            "input_stats": {},
            "probability_bands": {},
            "gpt_review": {"available": False},
            "output_sheets": {}
        }
        
        # Load and process input data
        dupes, cleaned = self._load_data(dupes_path, cleaned_path)
        
        # Calculate input statistics
        self._calculate_input_stats(dupes, cleaned)
        
        # Prepare merged data with probability bands
        merged, prob_bands = self._prepare_merged_data(dupes, cleaned)
        
        # Process GPT review if available
        gpt_flat = self._process_gpt_review(gpt_review_path) if gpt_review_path else []
        
        # Generate Excel report
        self._generate_excel_report(prob_bands, gpt_flat, report_path)
        
        # Update output sheet statistics
        self._update_output_stats(prob_bands, gpt_flat)
        
        return self.report_stats
    
    def _load_data(self, dupes_path: str, cleaned_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Load duplicate pairs and cleaned records data."""
        dupes = pd.read_csv(dupes_path)
        cleaned = pd.read_csv(cleaned_path).set_index("record_id")
        return dupes, cleaned
    
    def _calculate_input_stats(self, dupes: pd.DataFrame, cleaned: pd.DataFrame) -> None:
        """Calculate and store input data statistics."""
        self.report_stats["input_stats"] = {
            "total_pairs": len(dupes),
            "unique_records": len(cleaned),
            "mean_probability": float(dupes["prob"].mean()),
            "probability_quantiles": {
                "25%": float(dupes["prob"].quantile(0.25)),
                "50%": float(dupes["prob"].quantile(0.50)),
                "75%": float(dupes["prob"].quantile(0.75)),
                "90%": float(dupes["prob"].quantile(0.90))
            }
        }
    
    def _prepare_merged_data(
        self, 
        dupes: pd.DataFrame, 
        cleaned: pd.DataFrame
    ) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """Prepare merged data with probability bands."""
        # Combine duplicate pairs with full record details from both sides
        left = cleaned.loc[dupes["record_id_1"]].add_suffix("_1").reset_index()
        right = cleaned.loc[dupes["record_id_2"]].add_suffix("_2").reset_index()
        merged = pd.concat([dupes.reset_index(drop=True), left, right], axis=1)
        
        # Create probability bands
        prob_bands = {
            "high_confidence": merged[merged["prob"] >= 0.9],
            "manual_review": merged[(merged["prob"] >= 0.6) & (merged["prob"] < 0.9)],
            "low_confidence": merged[merged["prob"] < 0.6]
        }
        
        # Track probability band statistics
        self.report_stats["probability_bands"] = {
            band: {
                "count": len(df),
                "mean_probability": float(df["prob"].mean()) if len(df) > 0 else 0.0
            }
            for band, df in prob_bands.items()
        }
        
        return merged, prob_bands
    
    def _process_gpt_review(self, gpt_review_path: str) -> List[Dict[str, Any]]:
        """Process GPT review results if available."""
        gpt_flat = []
        
        if not os.path.exists(gpt_review_path):
            return gpt_flat
        
        try:
            with open(gpt_review_path, "r", encoding="utf-8") as f:
                gpt_review = json.load(f)
            
            # Track GPT review statistics
            total_clusters = len(gpt_review)
            processed_clusters = 0
            total_groups = 0
            total_records = 0
            confidence_sum = 0.0
            
            for entry in gpt_review:
                cluster_id = entry.get("cluster_id")
                record_ids = entry.get("record_ids", [])
                canonical_groups = entry.get("canonical_groups", [])
                
                # Only process clusters with non-empty canonical_groups
                if canonical_groups and isinstance(canonical_groups, list) and len(canonical_groups) > 0:
                    processed_clusters += 1
                    total_groups += len(canonical_groups)
                    
                    for group in canonical_groups:
                        # Required fields
                        primary_organization = group.get("primary_organization")
                        canonical_domains = group.get("canonical_domains", [])
                        canonical_record_ids = group.get("record_ids", [])
                        confidence = group.get("confidence", 0.0)
                        
                        if primary_organization and canonical_record_ids:
                            total_records += len(canonical_record_ids)
                            confidence_sum += confidence
                            gpt_flat.append({
                                "cluster_id": cluster_id,
                                "primary_organization": primary_organization,
                                "canonical_domains": canonical_domains,
                                "record_ids": canonical_record_ids,
                                "confidence": confidence
                            })
            
            # Update GPT review stats
            self.report_stats["gpt_review"] = {
                "available": True,
                "total_clusters": total_clusters,
                "processed_clusters": processed_clusters,
                "total_groups": total_groups,
                "total_records_processed": total_records,
                "mean_confidence": float(confidence_sum / len(gpt_flat)) if gpt_flat else 0.0,
                "records_per_group": float(total_records / total_groups) if total_groups > 0 else 0.0
            }
            
        except (json.JSONDecodeError, FileNotFoundError, KeyError) as e:
            # If GPT review processing fails, just mark as unavailable
            self.report_stats["gpt_review"] = {"available": False, "error": str(e)}
        
        return gpt_flat
    
    def _generate_excel_report(
        self, 
        prob_bands: Dict[str, pd.DataFrame], 
        gpt_flat: List[Dict[str, Any]], 
        report_path: str
    ) -> None:
        """Generate the Excel report with multiple sheets."""
        os.makedirs(os.path.dirname(report_path), exist_ok=True)
        
        with ExcelWriter(report_path, engine="openpyxl") as writer:
            # Write probability band sheets
            prob_bands["high_confidence"].to_excel(writer, sheet_name="high_confidence", index=False)
            prob_bands["manual_review"].to_excel(writer, sheet_name="manual_review", index=False)
            
            # Write GPT review sheet if available
            if gpt_flat:
                pd.DataFrame(gpt_flat).to_excel(writer, sheet_name="gpt_review", index=False)
            
            # Apply formatting to manual review sheet
            self._apply_excel_formatting(writer, prob_bands["manual_review"])
    
    def _apply_excel_formatting(self, writer: ExcelWriter, manual_review_df: pd.DataFrame) -> None:
        """Apply formatting to Excel sheets for better readability."""
        try:
            workbook = writer.book
            
            # Format manual review sheet with highlighting
            if "manual_review" in writer.sheets:
                review_ws = writer.sheets["manual_review"]
                fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                
                # Find probability column
                prob_idx = manual_review_df.columns.get_loc("prob")
                if isinstance(prob_idx, int):
                    score_col = prob_idx + 1  # Excel columns are 1-indexed
                    
                    # Apply highlighting to probability column
                    for row in review_ws.iter_rows(
                        min_row=2, max_row=review_ws.max_row, min_col=score_col, max_col=score_col
                    ):
                        for cell in row:
                            cell.fill = fill
        except Exception:
            # If formatting fails, continue without it
            pass
    
    def _update_output_stats(self, prob_bands: Dict[str, pd.DataFrame], gpt_flat: List[Dict[str, Any]]) -> None:
        """Update output sheet statistics."""
        self.report_stats["output_sheets"] = {
            "high_confidence": len(prob_bands["high_confidence"]),
            "manual_review": len(prob_bands["manual_review"]),
        }
        
        if gpt_flat:
            self.report_stats["output_sheets"]["gpt_review"] = len(gpt_flat)
    
    def get_summary_stats(self) -> Dict[str, Any]:
        """Get simplified summary statistics for logging."""
        if not self.report_stats:
            return {}
        
        return {
            "total_pairs": self.report_stats.get("input_stats", {}).get("total_pairs", 0),
            "high_confidence_pairs": self.report_stats.get("output_sheets", {}).get("high_confidence", 0),
            "manual_review_pairs": self.report_stats.get("output_sheets", {}).get("manual_review", 0),
            "gpt_review_available": self.report_stats.get("gpt_review", {}).get("available", False),
            "mean_probability": self.report_stats.get("input_stats", {}).get("mean_probability", 0.0)
        }
